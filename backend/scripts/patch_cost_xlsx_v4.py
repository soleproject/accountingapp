"""
Patch SmartBooks_Cost_Model.xlsx (v4):
  1. Update ARPU references to the actual product pricing tiers
     ($38 / $79 / $95 / $149) where applicable
  2. Add a "Security" tab: WAF, SSO, Secret Mgmt, Bug Bounty, DAST,
     Endpoint Security, IR retainer, scaled by user tier
  3. Add a "Hire vs Outsource" tab: what the 2-dev team + Emergent
     handles in-house vs. what to outsource, with hire cost and
     the user/revenue trigger for when the hire becomes necessary

Idempotent — safe to run multiple times.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = Path("/app/frontend/public/downloads/SmartBooks_Cost_Model.xlsx")

# ----------------------------------------------------------------------
# Styles
# ----------------------------------------------------------------------
NAVY = "0F172A"
CYAN = "0891B2"
SLATE_50 = "F8FAFC"
SLATE_100 = "F1F5F9"
AMBER = "F59E0B"
EMERALD = "059669"

TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
H2 = Font(name="Calibri", size=12, bold=True, color=NAVY)
H_TABLE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10, color=NAVY)
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=NAVY)
MUTED = Font(name="Calibri", size=9, color="64748B", italic=True)

FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_SUB = PatternFill("solid", fgColor=SLATE_100)
FILL_TOTAL = PatternFill("solid", fgColor="E0F2FE")
FILL_ACCENT = PatternFill("solid", fgColor="FEF3C7")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(style="thin", color="CBD5E1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_row(ws, row, values, font=BODY, fill=None, align=None, border=None,
            number_formats=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
        if border: c.border = border
        if number_formats and i - 1 < len(number_formats) and number_formats[i - 1]:
            c.number_format = number_formats[i - 1]


def widen(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------------------------------
# 1. SECURITY tab
# ----------------------------------------------------------------------
def build_security_tab(wb):
    name = "8. Security"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    widen(ws, [34, 42, 12, 12, 12, 12, 12, 12])

    ws.cell(row=1, column=1, value="Operational Security — Line-Item Costs by User Tier").font = TITLE
    ws.merge_cells("A1:H1")

    ws.cell(row=2, column=1, value=(
        "Ship what a real accounting SaaS needs. Bank data + PII means "
        "security is not optional. Numbers are $/mo unless noted."
    )).font = MUTED
    ws.merge_cells("A2:H2")

    # Header row
    headers = ["Category", "Line item", "Pre-launch", "500 low", "500 high",
               "1,500 low", "1,500 high", "3,000+"]
    set_row(ws, 4, headers, font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)

    rows = [
        # (Category, Item, pre, 500l, 500h, 1.5kl, 1.5kh, 3k)
        ("Perimeter",         "Cloudflare WAF + DDoS + Bot Mgmt",           25, 60, 120, 200, 300, 500),
        (None,                 "Cloudflare Zero Trust (team VPN)",           0,  0,  50,  100, 150, 250),
        ("Secrets Mgmt",      "Doppler / Infisical (env + rotation)",       0,  25, 50,  60,  120, 240),
        (None,                 "AWS/Google Secret Manager (per-svc keys)",  10, 20, 40,  60,  100, 200),
        ("SSO / IAM",         "SSO (SAML/OIDC) for Enterprise plan",        0,  0,  0,   0,   250, 500),
        (None,                 "1Password Business (per seat, team of 6)",  48, 48, 48,  96,  96,  144),
        (None,                 "MFA enforcement (Auth0/WorkOS add-on)",     0,  50, 100, 150, 300, 500),
        ("App Security",      "Snyk / Semgrep (SAST + dep scanning)",       0,  30, 60,  100, 200, 400),
        (None,                 "GitHub Advanced Security",                  0,  0,  49,  49,  147, 294),
        (None,                 "DAST (StackHawk / Detectify weekly scan)",  0,  0,  100, 250, 400, 700),
        ("Bug Bounty",        "HackerOne / Intigriti public program",       0,  0,  0,   500, 800, 1500),
        (None,                 "Bounty pool (avg payouts, budgeted)",       0,  0,  0,   300, 600, 1200),
        ("Endpoint",          "MDM (Kandji/Jamf, per device)",              0,  60, 60,  180, 240, 480),
        (None,                 "EDR (CrowdStrike/SentinelOne, per seat)",   0,  36, 36,  108, 144, 288),
        ("Logging & Alerts",  "SIEM (Panther / Datadog Cloud SIEM)",        0,  0,  100, 200, 500, 1000),
        (None,                 "Audit-log retention (S3 Glacier)",          5,  10, 20,  40,  80,  160),
        ("Backup / DR",       "Cross-region MongoDB backup + PITR",         20, 40, 80,  120, 200, 400),
        (None,                 "Immutable off-site backup (Wasabi/S3 lock)",10, 20, 40,  60,  120, 240),
        ("Compliance Ops",    "Vanta / Drata (SOC 2 automation)",           0,  0,  0,   800, 1250, 1250),
        (None,                 "Annual pen-test (amortised /mo)",           0,  0,  0,   500, 750, 1000),
        (None,                 "IR retainer (Kroll / Mandiant)",            0,  0,  0,   0,   400, 800),
        ("Fraud / Abuse",     "reCAPTCHA Enterprise + IP intel",            0,  20, 40,  80,  150, 300),
        (None,                 "Rate-limit + abuse mgmt (Kong/Cloudflare)", 0,  0,  30,  60,  120, 240),
    ]

    current_cat = ""
    r = 5
    for cat, item, *vals in rows:
        if cat and cat != current_cat:
            set_row(ws, r, [cat] + [""] * 7, font=BODY_BOLD, fill=FILL_SUB, align=LEFT, border=BOX)
            current_cat = cat
            r += 1
        set_row(ws, r,
                ["" if cat is None else "", item, *vals],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", "$#,##0", "$#,##0",
                                "$#,##0", "$#,##0", "$#,##0"])
        r += 1

    # Totals row
    total_row = r + 1
    ws.cell(row=total_row, column=1, value="TOTAL SECURITY / MONTH").font = BODY_BOLD
    ws.cell(row=total_row, column=1).fill = FILL_TOTAL
    ws.cell(row=total_row, column=1).border = BOX
    ws.cell(row=total_row, column=2).fill = FILL_TOTAL
    ws.cell(row=total_row, column=2).border = BOX

    for col in range(3, 9):
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_letter}5:{col_letter}{r - 1})")
        c.font = BODY_BOLD
        c.fill = FILL_TOTAL
        c.number_format = "$#,##0"
        c.border = BOX
        c.alignment = RIGHT

    # Notes
    n = total_row + 2
    ws.cell(row=n, column=1, value="Notes & phasing").font = H2
    notes = [
        "• Pre-launch = what you must have on Day 1 (Cloudflare, secrets, MFA, 1Password, encrypted backups).",
        "• 500-user tier = layer in Snyk, MDM, EDR, basic SIEM. This is the SOC-2 Type I readiness moment.",
        "• 1,500-user tier = SOC 2 automation (Vanta), pen-test, DAST, bug bounty program launches.",
        "• 3,000+ = mature program — SIEM w/ retention, IR retainer, GitHub Advanced Security, dedicated fraud tooling.",
        "• Numbers exclude Vanta itself in the 500 column (still in Compliance + Econ tab); avoid double-counting.",
        "• All prices are conservative-mid list; real spend often 10-20% lower after annual commits.",
    ]
    for i, note in enumerate(notes):
        c = ws.cell(row=n + 1 + i, column=1, value=note)
        c.font = BODY
        c.alignment = LEFT
        ws.merge_cells(start_row=n + 1 + i, start_column=1, end_row=n + 1 + i, end_column=8)

    ws.freeze_panes = "C5"


# ----------------------------------------------------------------------
# 2. HIRE vs OUTSOURCE tab
# ----------------------------------------------------------------------
def build_hire_outsource_tab(wb):
    name = "9. Hire vs Outsource"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    widen(ws, [26, 22, 34, 22, 20, 16, 22])

    ws.cell(row=1, column=1, value="Hire vs Outsource — What we build vs. what we contract").font = TITLE
    ws.merge_cells("A1:G1")

    ws.cell(row=2, column=1, value=(
        "Reality check: 2-founder team + Emergent dev credits. This tab shows "
        "what's genuinely covered by that setup, what should be outsourced "
        "now (fractional / contract), and when a full-time hire becomes "
        "necessary — with the revenue or user trigger."
    )).font = MUTED
    ws.merge_cells("A2:G2")

    # ---- Table 1: What we handle IN-HOUSE today ----
    ws.cell(row=4, column=1, value="A. What the current team (2 founders + Emergent) handles").font = H2
    ws.merge_cells("A4:G4")

    headers = ["Function", "Owner today", "Coverage", "Risk if we stay solo",
               "Backup plan", "Confidence", "Notes"]
    set_row(ws, 5, headers, font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)

    inhouse = [
        # Function, Owner, Coverage, Risk, Backup, Confidence, Notes
        ("Backend + API dev",       "You + Emergent",   "Full feature build; hot-reload FastAPI + Mongo",
         "Bus factor of 1",             "Emergent contract dev on retainer",  "High",
         "Emergent dev credits cover surges"),
        ("Frontend (React)",        "You + Emergent",   "React + Tailwind + shadcn",
         "UI polish behind roadmap",   "Fractional design contractor",       "High",
         "Design system already in place"),
        ("Product management",      "You",              "Roadmap, feature spec, user research",
         "You become the bottleneck",  "Hire PM at ~1,500 users",            "High",
         "You know the accounting domain cold"),
        ("Basic DevOps",            "You + Emergent",   "Railway + Atlas + supervisor",
         "3am pager on outages",       "Ops-as-service (Cloudops)",          "Medium",
         "Managed platform absorbs most work"),
        ("QA — smoke + unit",       "You + subagents",  "pytest + testing-agent",
         "Coverage gaps around edge cases", "Contract QA at 750 users",     "Medium",
         "Test agents good, human QA still needed"),
        ("Sales — founder-led",     "You",              "Direct outbound to first 100 accountants",
         "Doesn't scale past ~250 clients", "Hire sales lead at $30k MRR",  "High",
         "Best channel until Series A"),
        ("Content + copywriting",   "You + LLM",        "Blog, pitch decks, docs",
         "Voice / brand consistency",  "Fractional content writer",         "Medium",
         "LLM does 80%; human editor for polish"),
        ("Bookkeeping (of SmartBooks)", "You",         "Ironically, you can do your own books",
         "Time drain when scaling",    "Outsource at Series A",              "High",
         "Eat your own dogfood"),
    ]

    r = 6
    for row in inhouse:
        set_row(ws, r, list(row), font=BODY, align=LEFT, border=BOX)
        r += 1

    # ---- Table 2: OUTSOURCE now (fractional / contract) ----
    r += 2
    ws.cell(row=r, column=1, value="B. Outsource / contract NOW (fractional, not full-time)").font = H2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    headers2 = ["Function", "Provider type", "Why outsource",
                "Cost range ($/mo)", "When to start", "Full-time trigger", "Notes"]
    set_row(ws, r, headers2, font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)
    r += 1

    outsource = [
        ("Legal / TOS + Privacy",       "Startup law firm (Cooley / Gunderson) or LegalZoom+lawyer review",
         "Bank data + PII = real regulatory exposure",
         "$500–$1,500 flat / project", "Day 1", "General Counsel @ $8M ARR",
         "One-time TOS + Privacy $3–5k, then retainer"),

        ("Accounting / Tax (yours)",    "Fractional accountant / bookkeeper",
         "You're building an accounting product — filing DE + UK taxes wrong is a bad look",
         "$300–$800", "Day 1",  "In-house CFO @ $2M ARR",
         "Especially UK VAT + HMRC filings"),

        ("Compliance / SOC 2",         "Vanta + independent auditor",
         "You can't self-attest — auditor must be external",
         "$800–$1,500", "1,000 users", "Full-time Compliance @ 5,000 users",
         "Vanta handles evidence; auditor signs"),

        ("Design (brand + product)",   "Fractional designer / studio",
         "Founders can ship UI but marketing site + brand consistency needs a designer",
         "$1,500–$4,000", "Pre-launch marketing site", "Full-time Designer @ 2,000 users",
         "Especially for pitch decks + demos"),

        ("Cybersecurity audit",        "Boutique pen-test firm (Bishop Fox, Cure53)",
         "You can't SOC 2 without an external pen-test",
         "$500–$1,000 (amortised)", "1,500 users", "In-house Security Eng @ 5,000 users",
         "Annual test = $6–12k one-shot"),

        ("Customer support / SDR",     "Fractional VA / offshore SDR",
         "Save founder time on demos, scheduling, tier-1 questions",
         "$800–$2,500", "First paid signups", "In-house CS @ $20k MRR",
         "Time zone bonus for UK coverage"),

        ("Marketing / SEO / paid",     "Fractional CMO or agency",
         "You need SEO from Day 1 but shouldn't hire full-time yet",
         "$2,000–$5,000", "Post-launch", "Head of Growth @ $50k MRR",
         "Content-led compounds slowly, start early"),

        ("Payroll / HR",               "Gusto / Deel / Rippling",
         "SaaS handles this; DIY is a mistake past 3 people",
         "$40–$120 + $10/employee", "First hire", "HR person @ 20+ headcount",
         "Deel if hiring UK/EU contractors"),

        ("Recruiting",                 "Contingent recruiter",
         "Only pay on hire; save you weeks of screening",
         "15–20% first-year salary /hire", "Second hire",
         "In-house recruiter @ 25+ headcount",
         "Skip for junior hires — use LinkedIn"),
    ]
    for row in outsource:
        set_row(ws, r, list(row), font=BODY, align=LEFT, border=BOX)
        r += 1

    # ---- Table 3: When to HIRE full-time ----
    r += 2
    ws.cell(row=r, column=1, value="C. Full-time hires — role, US mid salary, and trigger").font = H2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    headers3 = ["Role", "Hire order", "Base salary ($k)",
                "Loaded cost ($k/yr)", "Trigger", "Rationale", "Alt: keep fractional?"]
    set_row(ws, r, headers3, font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)
    r += 1

    hires = [
        ("Senior Full-Stack Eng #1", "1st hire", 150, 195, "$25k MRR OR launch of UK",
         "Second builder to remove bus factor + accelerate roadmap",
         "No — need permanent code owner"),

        ("Customer Success Lead",    "2nd hire", 75,  100, "$40k MRR (~1,000 paid users)",
         "Onboarding + retention on autopilot",
         "Fractional VA until $20k MRR, then must hire"),

        ("Head of Growth / Marketing", "3rd hire", 130, 170, "$60k MRR",
         "Move from founder-led to demand-gen engine",
         "Yes — fractional CMO can run to $100k MRR"),

        ("Senior Full-Stack Eng #2", "4th hire", 150, 195, "1,500 users OR $60k MRR",
         "Split frontend / backend ownership; on-call rotation",
         "No — need full-time"),

        ("Head of Sales / AE",       "5th hire", 120, 160, "$80k MRR + $150k+ ACVs",
         "Enterprise/Partner tier deals need dedicated closer",
         "Yes — fractional sales advisor until then"),

        ("DevOps / Platform Eng",    "6th hire", 160, 210, "2,000 users OR SOC 2 Type II",
         "Uptime, SRE, IaC — bank-data SaaS needs this",
         "Fractional Cloudops fine to ~1,500 users"),

        ("Compliance Lead",          "7th hire", 130, 170, "SOC 2 Type II + ISO 27001 push",
         "Full-time evidence + vendor mgmt + audit prep",
         "Vanta + fractional to 5,000 users"),

        ("Design Lead",              "8th hire", 130, 170, "2,000 users",
         "Product design consistency; brand ownership",
         "Fractional works longer than most think"),

        ("Data / Analytics Eng",     "9th hire", 145, 190, "3,000 users",
         "Feature analytics, ML on categorisation, customer insights",
         "Contract data analyst can bridge"),

        ("General Counsel",          "10th hire", 175, 225, "$8M ARR",
         "Contracts, IP, partnerships, disputes",
         "Yes — outside counsel scales far"),

        ("Head of Finance / CFO",    "11th hire", 200, 260, "Series A or $5M ARR",
         "Board reporting, forecasting, fundraising, taxes",
         "Fractional CFO to Series A"),
    ]
    for row in hires:
        vals = list(row)
        set_row(ws, r, vals, font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", "$#,##0", None, None, None])
        r += 1

    # ---- Table 4: Rule of thumb summary ----
    r += 2
    ws.cell(row=r, column=1, value="D. Rules of thumb").font = H2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    rules = [
        "• Loaded cost ≈ base × 1.3 (US) or × 1.15 (UK/EU contractor via Deel).",
        "• Never hire full-time for something you can't keep busy 60%+ of the week — use fractional.",
        "• First 5 hires shape company culture; over-index on hiring signal, under-index on comp comparisons.",
        "• Every role above has a fractional or contract alternative — always try that first for one quarter.",
        "• Rule of engineering headcount: 1 engineer per ~$500k–$1M ARR in accounting SaaS (higher because of compliance overhead).",
        "• Rule for Bay Area vs. remote-first: Remote-first saves ~25% on comp with similar quality if you interview well.",
    ]
    for rule in rules:
        c = ws.cell(row=r, column=1, value=rule)
        c.font = BODY
        c.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1

    ws.freeze_panes = "A6"


# ----------------------------------------------------------------------
# 3. Update PRICING tiers where applicable
# ----------------------------------------------------------------------
def update_pricing_tiers(wb):
    """
    Update ARPU references to SmartBooks' actual product pricing:
      • Solo/Starter    $38/mo
      • Standard        $79/mo
      • Pro             $95/mo
      • Enterprise      $149/mo
    Blended ARPU used for MRR forecasts assumes: 40% Solo, 35% Standard,
    15% Pro, 10% Enterprise  →  0.40*38 + 0.35*79 + 0.15*95 + 0.10*149
                              = 15.20 + 27.65 + 14.25 + 14.90 = 72.00
    """
    blended = 0.40 * 38 + 0.35 * 79 + 0.15 * 95 + 0.10 * 149  # ~72.00

    # 3a. Summary tab — add pricing tier block below existing content
    ws = wb["1. Summary"]
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="PRODUCT PRICING TIERS (Feb 2026)").font = H2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    set_row(ws, r,
            ["Tier", "Price / mo", "Target segment", "Assumed mix %", "Blended contribution", ""],
            font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)
    r += 1
    tiers = [
        ("Solo / Starter", 38,  "Sole trader / 1 client", 0.40),
        ("Standard",       79,  "Bookkeeper w/ up to 5 clients", 0.35),
        ("Pro",            95,  "Firm w/ 10-30 clients", 0.15),
        ("Enterprise",     149, "Partner / 50+ clients / SSO", 0.10),
    ]
    for name_, price, segment, mix in tiers:
        set_row(ws, r,
                [name_, price, segment, mix, price * mix, ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, "0%", "$#,##0.00", None])
        r += 1
    set_row(ws, r,
            ["Blended ARPU", blended, "Used for revenue forecasts below", 1.0, blended, ""],
            font=BODY_BOLD, fill=FILL_TOTAL, align=LEFT, border=BOX,
            number_formats=[None, "$#,##0.00", None, "0%", "$#,##0.00", None])
    r += 2
    ws.cell(row=r, column=1,
            value=("Assumed mix is intentionally conservative — early cohort tends to "
                   "over-index on Standard + Pro; blended ARPU will land higher as "
                   "Enterprise deals close.")).font = MUTED
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    # 3b. 6-Mo Cash Flow — swap $35 ARPU → blended $72
    ws2 = wb["3. 6-Mo Cash Flow"]
    # Row 15 currently: "REVENUE (assumes UK launch month 3, ARPU $35)"
    for row in range(1, ws2.max_row + 1):
        v = ws2.cell(row=row, column=1).value
        if isinstance(v, str) and "ARPU $35" in v:
            ws2.cell(row=row, column=1,
                     value=v.replace("ARPU $35", f"blended ARPU ${blended:.0f}"))
        if isinstance(v, str) and v.startswith("MRR @ $35 ARPU"):
            ws2.cell(row=row, column=1, value=f"MRR @ blended ${blended:.0f} ARPU")
            # Recompute row values using paid users row (row above)
            paid_row = row - 1
            for col in range(2, 8):
                letter = get_column_letter(col)
                ws2.cell(row=row, column=col,
                         value=f"={letter}{paid_row}*{blended:.2f}")
                ws2.cell(row=row, column=col).number_format = "$#,##0"
            # 6-mo total column (col 8) keep as SUM
            ws2.cell(row=row, column=8, value=f"=SUM(B{row}:G{row})").number_format = "$#,##0"

    # 3c. Compliance + Econ tab — update Unit Economics block ARPU references
    ws3 = wb["7. Compliance + Econ"]
    # Replace hardcoded ARPU rows with tier-based rows
    for row in range(1, ws3.max_row + 1):
        v = ws3.cell(row=row, column=1).value
        if isinstance(v, str) and v.startswith("ARPU @ break-even"):
            ws3.cell(row=row, column=1, value="ARPU @ blended tier mix ($72)")
            ws3.cell(row=row, column=2, value=blended)
            ws3.cell(row=row, column=3, value=blended)
            ws3.cell(row=row, column=4, value=blended)
            for c in range(2, 5):
                ws3.cell(row=row, column=c).number_format = "$#,##0.00"
        if isinstance(v, str) and v.startswith("ARPU @ healthy"):
            ws3.cell(row=row, column=1, value="ARPU @ Enterprise-heavy mix ($95)")
            ws3.cell(row=row, column=2, value=95)
            ws3.cell(row=row, column=3, value=95)
            ws3.cell(row=row, column=4, value=95)
            for c in range(2, 5):
                ws3.cell(row=row, column=c).number_format = "$#,##0"
        if isinstance(v, str) and v.startswith("MRR @ 4x ARPU"):
            ws3.cell(row=row, column=1, value=f"MRR @ blended ${blended:.0f} ARPU")
            ws3.cell(row=row, column=2, value=f"=500*{blended:.2f}")
            ws3.cell(row=row, column=3, value=f"=1500*{blended:.2f}")
            ws3.cell(row=row, column=4, value=f"=3000*{blended:.2f}")
            for c in range(2, 5):
                ws3.cell(row=row, column=c).number_format = "$#,##0"
        if isinstance(v, str) and v.startswith("Gross margin @ $35 ARPU"):
            ws3.cell(row=row, column=1,
                     value=f"Gross margin @ blended ${blended:.0f} ARPU")
        if isinstance(v, str) and v.startswith("Gross margin @ $50 ARPU"):
            ws3.cell(row=row, column=1, value="Gross margin @ $95 ARPU (Pro heavy)")


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main():
    assert XLSX.exists(), f"Missing {XLSX}"
    wb = load_workbook(XLSX)
    update_pricing_tiers(wb)
    build_security_tab(wb)
    build_hire_outsource_tab(wb)
    wb.save(XLSX)
    print(f"Patched OK: {XLSX}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
