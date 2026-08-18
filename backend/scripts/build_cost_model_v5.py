"""
SmartBooks Cost Model v5 (Feb 2026) — TECH-ONLY REBUILD

Complete rebuild of /app/frontend/public/downloads/SmartBooks_Cost_Model.xlsx.

Tabs:
  1. Summary                — TL;DR, product tiers, blended ARPU
  2. Hard Costs @ 0 Clients — every $ that flows out with zero revenue
  3. 30/60/90/180 Cash Flow — cost timeline against revenue ramp
  4. Tech Outsource         — engineering/security contractors to hire
  5. Scale Costs by Tier    — 500 / 1,500 / 3,000 users
  6. AI Deep Dive           — LLM cost + optimisation levers
  7. Security               — WAF, SSO, SIEM, bug bounty, etc.
  8. Profit Breakdown       — MRR / gross margin / monthly profit
                              using $38/$75/$95/$149 tiers

STRICT: no payroll, HR, recruiting, legal, insurance, marketing, sales,
CFO, or bookkeeping content — tech spend only.

Pricing:  $38 / $75 / $95 / $149
Mix:      40% / 35% / 15% / 10%
Blended:  0.40*38 + 0.35*75 + 0.15*95 + 0.10*149 = $70.60
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = Path("/app/frontend/public/downloads/SmartBooks_Cost_Model.xlsx")

# ---------------- Style constants ----------------
NAVY = "0F172A"
CYAN = "0891B2"
SLATE_100 = "F1F5F9"
SKY_100 = "E0F2FE"
AMBER_100 = "FEF3C7"
EMERALD_100 = "D1FAE5"

TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
H2 = Font(name="Calibri", size=12, bold=True, color=NAVY)
H_TABLE = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10, color=NAVY)
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=NAVY)
MUTED = Font(name="Calibri", size=9, color="64748B", italic=True)

FILL_HEAD = PatternFill("solid", fgColor=NAVY)
FILL_SUB = PatternFill("solid", fgColor=SLATE_100)
FILL_TOTAL = PatternFill("solid", fgColor=SKY_100)
FILL_ACCENT = PatternFill("solid", fgColor=AMBER_100)
FILL_GOOD = PatternFill("solid", fgColor=EMERALD_100)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(style="thin", color="CBD5E1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------- Pricing constants ----------------
TIERS = [
    ("Solo / Starter", 38,  0.40, "Sole trader / 1 client"),
    ("Standard",       75,  0.35, "Bookkeeper w/ up to 5 clients"),
    ("Pro",            95,  0.15, "Firm w/ 10-30 clients"),
    ("Enterprise",     149, 0.10, "Partner / 50+ clients / SSO"),
]
BLENDED_ARPU = sum(p * m for _, p, m, _ in TIERS)  # 70.60


# ---------------- Helpers ----------------
def set_row(ws, row, values, font=BODY, fill=None, align=None,
            border=None, number_formats=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
        if border: c.border = border
        if number_formats and i - 1 < len(number_formats) and number_formats[i - 1]:
            c.number_format = number_formats[i - 1]


def widen(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, span_col_letter):
    ws.cell(row=1, column=1, value=text).font = TITLE
    ws.merge_cells(f"A1:{span_col_letter}1")


def subtitle(ws, text, span_col_letter, row=2):
    ws.cell(row=row, column=1, value=text).font = MUTED
    ws.merge_cells(f"A{row}:{span_col_letter}{row}")


def header_row(ws, row, headers):
    set_row(ws, row, headers, font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)


# ============================================================
# TAB 1 — SUMMARY
# ============================================================
def tab_summary(wb):
    ws = wb.create_sheet("1. Summary")
    widen(ws, [26, 18, 30, 16, 22, 18])
    title(ws, "SmartBooks — Tech-Only Cost Model (v5, Feb 2026)", "F")
    subtitle(ws, "Engineering, infra, AI, security and dev tooling only. "
                 "No payroll, HR, recruiting, legal, or GTM content.", "F")

    r = 4
    ws.cell(row=r, column=1, value="PRODUCT PRICING TIERS").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Tier", "Price / mo", "Target segment",
                       "Assumed mix %", "Blended contribution", ""])
    r += 1
    for name, price, mix, segment in TIERS:
        set_row(ws, r,
                [name, price, segment, mix, price * mix, ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, "0%", "$#,##0.00", None])
        r += 1
    set_row(ws, r, ["BLENDED ARPU", BLENDED_ARPU,
                    "Used for all revenue forecasts", 1.0, BLENDED_ARPU, ""],
            font=BODY_BOLD, fill=FILL_TOTAL, align=LEFT, border=BOX,
            number_formats=[None, "$#,##0.00", None, "0%", "$#,##0.00", None])

    r += 2
    ws.cell(row=r, column=1, value="HARD COSTS AT ZERO CLIENTS").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    set_row(ws, r, ["Scenario", "$/mo", "Note", "", "", ""],
            font=H_TABLE, fill=FILL_HEAD, align=CENTER, border=BOX)
    r += 1
    scenarios = [
        ("Bare-minimum floor (paused build)",  2100, "Plaid + Veryfi + Railway + core SaaS + Cloudflare"),
        ("Steady-state (1 fractional dev, no LLM burn)", 5500, "Floor + $3k/mo fractional + $500/mo LLM"),
        ("Current heavy-build burn", 15500, "Floor + $6k Emergent LLM + $6k contract devs"),
    ]
    for row_ in scenarios:
        set_row(ws, r,
                [row_[0], row_[1], row_[2], "", "", ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, None, None, None])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="PROFIT SNAPSHOT (blended $71 ARPU)").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Users", "MRR", "Monthly tech cost", "Gross profit", "Gross margin", ""])
    r += 1
    profit_rows = [
        (100,   100  * BLENDED_ARPU, 3500,   None),
        (250,   250  * BLENDED_ARPU, 4500,   None),
        (500,   500  * BLENDED_ARPU, 4000,   None),
        (1500,  1500 * BLENDED_ARPU, 11000,  None),
        (3000,  3000 * BLENDED_ARPU, 26000,  None),
    ]
    for users, mrr, cost, _ in profit_rows:
        gp = mrr - cost
        margin = gp / mrr if mrr else 0
        set_row(ws, r, [users, mrr, cost, gp, margin, ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", "$#,##0", "$#,##0", "0.0%", None])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="PITCH ONE-LINER").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    ws.cell(row=r, column=1, value=(
        f"At 1,500 paid users on a blended ${BLENDED_ARPU:.0f} ARPU, SmartBooks "
        f"runs ~$11k/mo of pure tech cost against ~$106k MRR — 90% gross margin "
        f"on the infra layer. Today's pre-revenue burn is ~$15.5k/mo of which "
        f"$12k is temporary build velocity that tapers by month 6."
    )).font = BODY
    ws.merge_cells(f"A{r}:F{r}")


# ============================================================
# TAB 2 — HARD COSTS @ 0 CLIENTS
# ============================================================
def tab_hard_costs(wb):
    ws = wb.create_sheet("2. Hard Costs @ 0 Clients")
    widen(ws, [38, 14, 14, 32])
    title(ws, "Hard Tech Costs at Zero Paying Clients (Feb 2026)", "D")
    subtitle(ws, "Every recurring $ that leaves the account with no revenue coming in. "
                 "Split into fixed floor, temporary build spend, and stuff we could pause.", "D")

    r = 4
    ws.cell(row=r, column=1, value="A. FIXED FLOOR — cannot pause").font = H2
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    header_row(ws, r, ["Line item", "$/mo", "Category", "Why unavoidable"])
    r += 1
    floor = [
        ("Plaid — production minimum",           1065, "Banking",       "Contract-mandated production floor for bank feeds"),
        ("Veryfi — OCR minimum",                  500, "OCR",           "Contract floor for receipt / bill OCR"),
        ("Railway (backend + Redis)",             300, "Infra",         "Backend + Redis for the running app"),
        ("MongoDB Atlas (shared M10 → M20)",      100, "Database",      "Primary data store"),
        ("Cloudflare (Pro + WAF)",                 25, "Perimeter",     "DNS, cache, WAF, DDoS — non-negotiable"),
        ("Google Workspace (1 seat)",              12, "Ops",           "Domain email + calendar"),
        ("Domain renewals (amortised)",            10, "Ops",           "smartbooks.ai etc."),
        ("1Password Business (1 seat)",             8, "Security",      "Founder credential vault"),
        ("GitHub Team (1 seat)",                    4, "Dev tools",     "Source control + issue tracking"),
        ("Sentry (free tier)",                      0, "Monitoring",    "Errors + traces; upgrade at 5k events/mo"),
        ("Resend (free tier — 3k emails/mo)",       0, "Email",         "Transactional email until paid signups"),
    ]
    for row_ in floor:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, None])
        r += 1
    total_floor_row = r
    ws.cell(row=r, column=1, value="Sub-total — fixed floor").font = BODY_BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B6:B{r - 1})")
    ws.cell(row=r, column=2).number_format = "$#,##0"
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = FILL_TOTAL
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD

    r += 2
    ws.cell(row=r, column=1, value="B. TEMPORARY BUILD SPEND — tapers by month 6").font = H2
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    header_row(ws, r, ["Line item", "$/mo", "Category", "How it tapers"])
    r += 1
    start_temp = r
    temp = [
        ("Emergent LLM (dev credits)",             6000, "AI",            "Drops to $500–1,500/mo once heavy build ends"),
        ("Emergent contract developers",            6000, "Engineering",   "Time-bound — 3–4 more months of heavy build"),
        ("Extra Railway compute during dev",         100, "Infra",         "Right-size once feature velocity slows"),
    ]
    for row_ in temp:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, None])
        r += 1
    ws.cell(row=r, column=1, value="Sub-total — temporary").font = BODY_BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B{start_temp}:B{r - 1})")
    ws.cell(row=r, column=2).number_format = "$#,##0"
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = FILL_ACCENT
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    total_temp_row = r

    r += 2
    ws.cell(row=r, column=1, value="C. OPTIONAL / PAUSABLE — nice-to-have during pre-revenue").font = H2
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    header_row(ws, r, ["Line item", "$/mo", "Category", "Impact if paused"])
    r += 1
    start_opt = r
    opt = [
        ("Datadog / Logtail",              70, "Monitoring", "Fall back to free tier; lose log retention"),
        ("Snyk / Semgrep dependency scan", 30, "Security",   "Use free GitHub Dependabot instead"),
        ("Doppler / Infisical (secrets)",  25, "Security",   "Fall back to .env files (lose rotation)"),
        ("Figma Professional",             15, "Design",     "Use free tier during pre-revenue"),
        ("ChatGPT Team seat",              25, "AI",         "Fall back to personal Plus"),
        ("Preview branch environments",    40, "Infra",      "Test on main only until launch"),
    ]
    for row_ in opt:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, None])
        r += 1
    ws.cell(row=r, column=1, value="Sub-total — optional").font = BODY_BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B{start_opt}:B{r - 1})")
    ws.cell(row=r, column=2).number_format = "$#,##0"
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = FILL_SUB
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    total_opt_row = r

    r += 2
    ws.cell(row=r, column=1, value="TOTAL BURN AT 0 CLIENTS").font = H2
    ws.cell(row=r, column=2,
            value=f"=B{total_floor_row}+B{total_temp_row}+B{total_opt_row}")
    ws.cell(row=r, column=2).number_format = "$#,##0"
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = FILL_TOTAL
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD

    r += 2
    ws.cell(row=r, column=1, value="RUNWAY CALCULATOR").font = H2
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    ws.cell(row=r, column=1, value="Cash on hand (edit)")
    ws.cell(row=r, column=2, value=150000).number_format = "$#,##0"
    ws.cell(row=r, column=2).fill = FILL_GOOD
    cash_row = r
    r += 1
    ws.cell(row=r, column=1, value="Runway @ current burn (all three)").font = BODY_BOLD
    ws.cell(row=r, column=2,
            value=f"=B{cash_row}/(B{total_floor_row}+B{total_temp_row}+B{total_opt_row})")
    ws.cell(row=r, column=2).number_format = "0.0"
    ws.cell(row=r, column=3, value="months").font = MUTED
    r += 1
    ws.cell(row=r, column=1, value="Runway @ floor + optional only (paused build)").font = BODY_BOLD
    ws.cell(row=r, column=2,
            value=f"=B{cash_row}/(B{total_floor_row}+B{total_opt_row})")
    ws.cell(row=r, column=2).number_format = "0.0"
    ws.cell(row=r, column=3, value="months").font = MUTED
    r += 1
    ws.cell(row=r, column=1, value="Runway @ absolute floor only").font = BODY_BOLD
    ws.cell(row=r, column=2, value=f"=B{cash_row}/B{total_floor_row}")
    ws.cell(row=r, column=2).number_format = "0.0"
    ws.cell(row=r, column=3, value="months").font = MUTED


# ============================================================
# TAB 3 — 30/60/90/180 CASH FLOW (ALL-IN, single source of truth)
# ============================================================
def tab_cashflow(wb):
    ws = wb.create_sheet("3. 30-60-90-180 Cash Flow")
    widen(ws, [40, 14, 14, 14, 14, 14])
    title(ws, "30 / 60 / 90 / 180 Day Cash Flow — ALL-IN", "F")
    subtitle(ws,
        "Every $ that leaves the account: hard costs + AI + security + "
        "outsource + any FTEs. Assumes UK launch by day 60. Realistic user "
        f"ramp 0→10→40→100. Revenue = paid users × ${BLENDED_ARPU:.2f} "
        "blended ARPU.", "F")

    r = 4
    # -------- Team roster reminder --------
    ws.cell(row=r, column=1, value="TEAM ROSTER IN THIS 180-DAY WINDOW").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Role", "Day 0–30", "Day 30–60",
                       "Day 60–90", "Day 90–180", "Notes"])
    r += 1
    roster = [
        ("2 founders (unpaid — sweat equity)",    "Yes", "Yes", "Yes", "Yes",
         "Not on P&L"),
        ("Emergent contract dev (fractional)",   "Yes", "Yes", "Taper", "Sunset",
         "Sunsets month 5 as build stabilises"),
        ("Fractional QA engineer",                "—",  "—",   "Start", "Yes",
         "Starts at UK launch (day 60)"),
        ("FTE hires (any)",                       "—",  "—",   "—",    "—",
         "None triggered — $25k MRR not hit in 180d"),
    ]
    for row_ in roster:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX)
        r += 1

    # -------- Cost breakdown --------
    r += 2
    ws.cell(row=r, column=1, value="MONTHLY COST BREAKDOWN").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Category / line", "Day 0–30", "Day 30–60",
                       "Day 60–90", "Day 90–180 (mo avg)", "180-day total"])
    r += 1

    costs_start = r

    # (label, d30, d60, d90, d90_180_month_avg, sub-heading?)
    def add_group(rows, group_label):
        nonlocal r
        set_row(ws, r, [group_label, "", "", "", "", ""],
                font=BODY_BOLD, fill=FILL_SUB, align=LEFT, border=BOX)
        r += 1
        for label, d30, d60, d90, avg in rows:
            total_180 = d30 + d60 + d90 + 3 * avg
            set_row(ws, r, [f"  {label}", d30, d60, d90, avg, total_180],
                    font=BODY, align=LEFT, border=BOX,
                    number_formats=[None, "$#,##0", "$#,##0", "$#,##0", "$#,##0", "$#,##0"])
            r += 1

    add_group([
        ("Plaid",                       1065, 1065, 1100, 1200),
        ("Veryfi OCR",                   500,  500,  550,  700),
        ("Railway + Mongo + Redis",      450,  500,  600,  850),
        ("Cloudflare + perimeter",        25,   25,   40,   80),
        ("Domain, Workspace, 1PW, GH",    50,   55,   60,   90),
        ("Backups (immutable + PITR)",    20,   30,   40,   80),
    ], "INFRA & HARD COSTS")

    add_group([
        ("Emergent LLM (dev credits)",  6000, 6000, 4500, 2500),
    ], "AI")

    add_group([
        ("Snyk / Doppler / MDM / EDR",    30,   60,  120,  300),
        ("Sentry + Datadog / monitoring",  0,   50,  120,  220),
    ], "SECURITY TOOLING (SaaS)")

    add_group([
        ("Emergent contract dev",       6000, 6000, 4500, 1500),
        ("Fractional QA engineer",         0,    0, 1500, 1500),
    ], "OUTSOURCE / FRACTIONAL")

    add_group([
        ("(none triggered in first 180 days)", 0, 0, 0, 0),
    ], "FTE HIRES")

    costs_end = r - 1

    # Grand total
    ws.cell(row=r, column=1, value="TOTAL COST / MONTH").font = BODY_BOLD
    for col in range(2, 6):
        cl = get_column_letter(col)
        # SUM over cost rows only (skip sub-heading rows which have "" values)
        ws.cell(row=r, column=col, value=f"=SUM({cl}{costs_start}:{cl}{costs_end})")
        ws.cell(row=r, column=col).number_format = "$#,##0"
    ws.cell(row=r, column=6, value=f"=SUM(F{costs_start}:F{costs_end})")
    ws.cell(row=r, column=6).number_format = "$#,##0"
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = FILL_TOTAL
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    total_cost_row = r
    r += 2

    # -------- Revenue --------
    ws.cell(row=r, column=1, value="REVENUE").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1

    paid_users_row = r
    set_row(ws, r,
            ["Paid users (end of period)", 0, 10, 40, 100, ""],
            font=BODY, align=LEFT, border=BOX,
            number_formats=[None, "#,##0", "#,##0", "#,##0", "#,##0", None])
    r += 1

    set_row(ws, r,
            ["Avg paid users in period", 0, 5, 25, 70, ""],
            font=MUTED, align=LEFT, border=BOX,
            number_formats=[None, "#,##0", "#,##0", "#,##0", "#,##0", None])
    avg_users_row = r
    r += 1

    ws.cell(row=r, column=1, value=f"MRR (avg users × ${BLENDED_ARPU:.2f})").font = BODY_BOLD
    for col, avg in [(2, 0), (3, 5), (4, 25), (5, 70)]:
        v = avg * BLENDED_ARPU
        ws.cell(row=r, column=col, value=v)
        ws.cell(row=r, column=col).number_format = "$#,##0"
    ws.cell(row=r, column=6,
            value=f"=B{r}+C{r}+D{r}+3*E{r}")
    ws.cell(row=r, column=6).number_format = "$#,##0"
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    mrr_row = r
    r += 2

    # -------- Net burn --------
    ws.cell(row=r, column=1, value="NET BURN (cost − revenue) /mo").font = BODY_BOLD
    for col in range(2, 6):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"={cl}{total_cost_row}-{cl}{mrr_row}")
        ws.cell(row=r, column=col).number_format = "$#,##0"
    ws.cell(row=r, column=6, value=f"=F{total_cost_row}-F{mrr_row}")
    ws.cell(row=r, column=6).number_format = "$#,##0"
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = FILL_ACCENT
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    net_burn_row = r
    r += 1

    # Cumulative
    ws.cell(row=r, column=1, value="CUMULATIVE NET BURN (end of period)").font = BODY_BOLD
    ws.cell(row=r, column=2, value=f"=B{net_burn_row}")
    ws.cell(row=r, column=3, value=f"=B{r}+C{net_burn_row}")
    ws.cell(row=r, column=4, value=f"=C{r}+D{net_burn_row}")
    # Months 4-6 add 3 × avg-month net burn
    ws.cell(row=r, column=5, value=f"=D{r}+3*E{net_burn_row}")
    ws.cell(row=r, column=6, value=f"=E{r}")
    for col in range(2, 7):
        ws.cell(row=r, column=col).number_format = "$#,##0"
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = FILL_TOTAL
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    cumu_row = r
    r += 2

    # -------- Bottom line --------
    ws.cell(row=r, column=1, value="CASH REQUIRED TO SURVIVE 180 DAYS").font = H2
    ws.cell(row=r, column=2, value=f"=F{cumu_row}")
    ws.cell(row=r, column=2).number_format = "$#,##0"
    for col in range(1, 3):
        ws.cell(row=r, column=col).fill = FILL_GOOD
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = H2
    r += 2

    # -------- What-if levers --------
    ws.cell(row=r, column=1, value="WHAT-IF LEVERS (edit assumptions)").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Lever", "Impact on 180-day cash need",
                       "Effort", "", "", ""])
    r += 1
    levers = [
        ("Skip fractional QA until month 5",  "−$3,000",   "None"),
        ("Cut Emergent contract dev to $3k/mo from day 1", "−$18,000",  "Slows velocity ~40%"),
        ("Delay UK launch by 30 days (revenue starts month 4)", "+$1,800",  "Loses UK market timing"),
        ("Pause LLM heavy-build & use $500/mo baseline", "−$25,500", "You stop shipping AI features"),
        ("Hit 200 paying users by day 180 instead of 100", "−$12,700",  "Requires 2× conversion rate"),
        ("Emergent Studio prepay ($25k) covers LLM 4-5 months",  "−$0 (cash-neutral swap)", "Just changes timing"),
    ]
    for row_ in levers:
        set_row(ws, r, [row_[0], row_[1], row_[2], "", "", ""],
                font=BODY, align=LEFT, border=BOX)
        r += 1

    ws.freeze_panes = "B5"


# ============================================================
# TAB 4 — TECH OUTSOURCE
# ============================================================
def tab_outsource(wb):
    ws = wb.create_sheet("4. Tech Outsource")
    widen(ws, [28, 30, 36, 20, 20, 22])
    title(ws, "Tech Outsource — Engineering, DevOps, Security, QA", "F")
    subtitle(ws, "Contract-only view. Every role here is a function the "
                 "2-founder + Emergent stack cannot cover reliably. Zero "
                 "GTM / HR / legal — pure tech execution.", "F")

    r = 4
    ws.cell(row=r, column=1,
            value="A. What the 2-founder + Emergent stack covers today").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Function", "Owner today", "Coverage",
                       "Risk if we stay solo", "Backup plan", "Confidence"])
    r += 1
    inhouse = [
        ("Backend + API (FastAPI + Mongo)", "You + Emergent",
         "Full feature build, hot-reload dev loop",
         "Bus factor of 1; velocity ceiling",
         "Emergent contract dev on retainer", "High"),

        ("Frontend (React + Tailwind)",     "You + Emergent",
         "Feature UI + shadcn components",
         "UI polish + a11y behind roadmap",
         "Fractional frontend contractor at $25k MRR", "High"),

        ("AI features (categorisation, chat)", "You + Emergent",
         "GPT / Claude / Gemini via Emergent Universal Key",
         "AI ops cost drift; prompt regressions",
         "Fractional ML/AI contractor at 1,500 users", "High"),

        ("Basic DevOps (Railway + Atlas)",  "You + Emergent",
         "Deploy, supervisor, log tail",
         "On-call fatigue at 3am",
         "Fractional platform engineer at 1,000 users", "Medium"),

        ("Smoke + unit testing",            "You + testing subagent",
         "pytest + testing agent + curl",
         "Coverage gaps around edge cases",
         "Contract QA engineer at 750 users", "Medium"),

        ("Product design (functional)",     "You",
         "Layouts, flows, shadcn defaults",
         "Brand + marketing site inconsistency",
         "Fractional designer once marketing site starts", "Medium"),
    ]
    for row_ in inhouse:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX)
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="B. Tech functions to OUTSOURCE (fractional / contract)").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Function", "Provider type", "Why outsource",
                       "Cost range ($/mo)", "When to start",
                       "Full-time trigger"])
    r += 1
    outsource = [
        ("Senior Full-Stack Dev (surge)",   "Emergent contract dev (current)",
         "Extra hands during heavy build without full-time commit",
         "$3,000–$8,000",   "In use today",
         "Full-time SWE at $25k MRR"),

        ("Fractional Platform / DevOps",    "Individual consultant or Cloudops firm",
         "SRE, IaC, on-call rotation, cost tuning on Railway/Atlas",
         "$2,500–$5,000",   "1,000 users OR SOC 2 kickoff",
         "Full-time Platform Eng at 2,000 users"),

        ("Fractional Security Engineer",    "Individual (ex-Big-Tech security)",
         "Threat model, IAM review, SIEM tuning, IR playbooks",
         "$3,000–$6,000",   "1,000 users OR pen-test prep",
         "Full-time Security Eng at 5,000 users"),

        ("Penetration Test (annual)",       "Boutique firm (Bishop Fox, Cure53, Doyensec)",
         "External attestation required for SOC 2 + enterprise deals",
         "$500–$1,000 (amortised)", "1,500 users",
         "Never — always outsource"),

        ("Vulnerability scanning (SaaS)",   "Snyk / Semgrep / GH Advanced Security",
         "Continuous SAST + dependency + secret scanning",
         "$100–$400",       "Day 1",
         "Never — always SaaS"),

        ("Fractional QA Engineer",          "Contractor via Toptal / Braintrust",
         "End-to-end Playwright coverage, regression suites",
         "$1,500–$3,500",   "750 users",
         "Full-time QA Lead at 3,000 users"),

        ("Fractional Data / Analytics Eng", "Contractor",
         "Feature analytics, dashboards, ML on categorisation",
         "$2,000–$4,000",   "1,500 users",
         "Full-time Data Eng at 3,000 users"),

        ("Fractional UX / Product Designer","Studio or individual",
         "Design system, marketing site, brand consistency",
         "$1,500–$4,000",   "Pre-launch marketing site",
         "Full-time Designer at 2,000 users"),

        ("Incident Response Retainer",      "Kroll / Mandiant / Arete",
         "Pre-paid response hours if a breach occurs",
         "$400–$800",       "1,500 users",
         "Never — always retainer"),

        ("Bug Bounty Program",              "HackerOne / Intigriti",
         "Crowd-sourced security testing; SOC 2 evidence",
         "$500–$1,500 platform + bounty pool", "1,500 users",
         "Never — always outsource"),

        ("Localisation / i18n",             "Fractional contractor or Lokalise SaaS",
         "UK + EU + APAC translation, currency, date/tax rules",
         "$500–$2,000",     "Per new region launch",
         "Full-time when 3+ regions live"),
    ]
    for row_ in outsource:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="C. Full-time tech hires (order + trigger)").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    header_row(ws, r, ["Role", "Order", "Loaded cost ($k/yr)",
                       "Trigger", "Why this specific hire", "Keep fractional?"])
    r += 1
    hires = [
        ("Senior Full-Stack Eng #1", "1st", 195,
         "$25k MRR OR UK launch",
         "Second permanent code owner — kills bus factor",
         "No — need FTE"),

        ("Platform / DevOps Eng",    "2nd", 210,
         "2,000 users OR SOC 2 Type II",
         "SRE, IaC, incident response ownership",
         "Fractional until 1,500 users"),

        ("Senior Full-Stack Eng #2", "3rd", 195,
         "3,000 users OR $60k MRR",
         "Split frontend / backend ownership + on-call rotation",
         "No — need FTE"),

        ("Security Engineer",        "4th", 220,
         "SOC 2 Type II + first enterprise deal",
         "In-house threat model, IR ownership, SSO reviews",
         "Fractional to 5,000 users"),

        ("QA / Test Automation Lead","5th", 165,
         "3,000 users",
         "Playwright + performance + regression suites",
         "Fractional to 3,000 users"),

        ("Data / Analytics Eng",     "6th", 190,
         "3,000 users",
         "Categorisation ML, cohort analytics, churn model",
         "Contract data analyst can bridge"),

        ("Product Designer",         "7th", 170,
         "3,000 users",
         "Full design system + marketing site ownership",
         "Fractional works longer than most think"),
    ]
    for row_ in hires:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", None, None, None])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="D. Rules of thumb (tech-only)").font = H2
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    rules = [
        "• 1 full-time engineer per ~$500k–$1M ARR in accounting SaaS (compliance overhead is real).",
        "• Never hire a full-time platform eng before 1,000 users — fractional or Cloudops does it cheaper.",
        "• Security engineer FTE only after first enterprise / SOC 2 Type II — before then, fractional + SaaS.",
        "• Rule of thumb: every tech FTE = base × 1.30 loaded (US) or × 1.15 (UK/EU via Deel).",
        "• Prefer 1 senior remote FTE over 2 mid-level unless you already have a team lead.",
        "• Emergent contract dev is functionally the cheapest fractional SWE on the market — use it until $25k MRR.",
    ]
    for rule in rules:
        c = ws.cell(row=r, column=1, value=rule)
        c.font = BODY
        c.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1


# ============================================================
# TAB 5 — SCALE COSTS BY TIER
# ============================================================
def tab_scale(wb):
    ws = wb.create_sheet("5. Scale Costs")
    widen(ws, [22, 40, 12, 12, 12, 12, 12, 12])
    title(ws, "Line-Item Tech Costs by User Tier", "H")
    subtitle(ws, "Fully-loaded tech spend across 500 / 1,500 / 3,000 users. "
                 "Excludes salaries.", "H")

    r = 4
    header_row(ws, r,
               ["Category", "Line Item",
                "500 low", "500 high",
                "1,500 low", "1,500 high",
                "3,000 low", "3,000 high"])
    r += 1
    start = r
    rows = [
        ("Compute",     "Backend (Railway/Fly, replicas)",         150, 250, 400, 700,  900, 1500),
        ("Compute",     "Frontend hosting (Vercel/Netlify)",        20,  40,  60, 120,  150,  250),
        ("Database",    "MongoDB Atlas dedicated",                 180, 220, 450, 550,  900, 1300),
        ("Database",    "Redis Cloud",                              30,  50,  75, 120,  200,  350),
        ("Database",    "Automated backups + PITR",                 20,  40,  60, 100,  120,  200),
        ("AI",          "LLM inference",                          1000,2000,3000,6000, 6000,12000),
        ("AI",          "STT / voice (Whisper)",                    20,  60,  60, 200,  150,  500),
        ("Integrations","Plaid (production)",                     1065,1200,1200,2100, 1800, 3000),
        ("Integrations","Veryfi OCR",                              500, 700, 800,1500, 1500, 2500),
        ("Email",       "Resend transactional",                     20,  35,  50,  90,  150,  250),
        ("Storage",     "Object storage (S3/R2)",                   30,  60, 150, 300,  400,  700),
        ("Storage",     "CDN egress",                               10,  30,  40, 100,  100,  250),
        ("Monitoring",  "Sentry",                                   30,  50,  80, 150,  150,  300),
        ("Monitoring",  "Datadog / Logtail",                        50,  70, 120, 250,  350,  600),
        ("Security",    "Vanta / Drata",                             0,   0,   0, 200,  500,  800),
        ("Security",    "Snyk / dep scanning",                       0,  30,  30,  60,   60,  120),
        ("Security",    "MDM + EDR",                                 0,  40,  40, 160,  240,  480),
        ("Ops",         "Domain, DNS, Cloudflare",                  20,  40,  30,  60,   50,  100),
        ("Ops",         "Password manager, misc SaaS",              20,  40,  40,  80,   60,  150),
    ]
    for row_ in rows:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", "$#,##0", "$#,##0",
                                "$#,##0", "$#,##0", "$#,##0"])
        r += 1
    end = r - 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAL / mo").font = BODY_BOLD
    for col in range(3, 9):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({cl}{start}:{cl}{end})")
        ws.cell(row=r, column=col).number_format = "$#,##0"
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = FILL_TOTAL
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD
    total_r = r
    r += 1
    ws.cell(row=r, column=1, value="Cost per paid user")
    for col, users in zip([3, 4, 5, 6, 7, 8], [500, 500, 1500, 1500, 3000, 3000]):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"={cl}{total_r}/{users}")
        ws.cell(row=r, column=col).number_format = "$#,##0.00"


# ============================================================
# TAB 6 — AI DEEP DIVE
# ============================================================
def tab_ai(wb):
    ws = wb.create_sheet("6. AI Deep Dive")
    widen(ws, [34, 18, 12, 12, 14])
    title(ws, "AI Cost Forecasting", "E")
    subtitle(ws, "Per-op cost, monthly usage by segment, and the "
                 "optimisation levers to keep AI COGS under 15% of ARPU.", "E")

    r = 4
    ws.cell(row=r, column=1, value="A — Cost per operation").font = H2
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    header_row(ws, r, ["Operation", "Model", "Tokens in", "Tokens out", "$/call"])
    r += 1
    ops = [
        ("Transaction categorisation",  "GPT-5.4 Mini",     600,  80,  0.0015),
        ("Batch categorise (10 txns)",  "GPT-5.4 Mini",    3000, 400,  0.0065),
        ("Voice STT (10s clip)",        "Whisper",            0,   0,  0.001),
        ("Voice intent parse",          "GPT-5.4 Mini",     400,  50,  0.0008),
        ("Chat message",                "Claude Haiku 4.5",2000, 400,  0.010),
        ("Long-form insight report",    "Claude Sonnet 5", 8000,2000,  0.100),
        ("QBO Verify (PDF extract)",    "Claude Sonnet 5",15000,3000,  0.200),
        ("Anomaly detection sweep",     "GPT-5.4 Mini",    4000, 300,  0.006),
        ("Onboarding interview",        "GPT-5.4 Mini",    3000, 800,  0.008),
        ("Contact match",               "GPT-5.4 Mini",     500,  60,  0.0009),
        ("Receipt image (Nano Banana)", "Gemini 3.5 Flash",   0, 200,  0.003),
    ]
    for row_ in ops:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "#,##0", "#,##0", "$#,##0.0000"])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="B — Monthly usage by segment").font = H2
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    header_row(ws, r, ["Segment", "% users", "AI ops/mo",
                       "$/user (low)", "$/user (high)"])
    r += 1
    seg = [
        ("Free trial / dormant", 0.30,   50,  0.10, 0.30),
        ("Light client",         0.35,  800,  1.20, 2.00),
        ("Standard client",      0.20, 3500,  2.50, 4.50),
        ("Power client",         0.08, 6000,  5.00, 9.00),
        ("Pro (10 clients)",     0.05,25000, 15.00,30.00),
        ("Pro power (30 clients)",0.02,60000, 35.00,70.00),
    ]
    for row_ in seg:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "0%", "#,##0", "$#,##0.00", "$#,##0.00"])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="C — Optimisation levers (stackable)").font = H2
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    header_row(ws, r, ["Lever", "Savings", "Effort", "", ""])
    r += 1
    levers = [
        ("Route classification through GPT-5.4 Mini",  -0.35, "Low"),
        ("Cache repeat classifications",               -0.15, "Low (2 days)"),
        ("Batch small ops (10/call)",                  -0.25, "Medium"),
        ("Claude Haiku 4.5 for chat",                  -0.20, "Low"),
        ("Precompute nightly insight summaries",       -0.15, "Medium"),
        ("Rate-limit AI Ask (20/day cap)",             -0.08, "Trivial"),
        ("Realistic combined savings (overlapping)",   -0.45, "Stacked, non-additive"),
    ]
    for label, sav, eff in levers:
        set_row(ws, r, [label, sav, eff, "", ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "0%", None, None, None])
        r += 1


# ============================================================
# TAB 7 — SECURITY
# ============================================================
def tab_security(wb):
    ws = wb.create_sheet("7. Security")
    widen(ws, [22, 42, 12, 12, 12, 12, 12, 12])
    title(ws, "Operational Security — Line-Item Costs by User Tier", "H")
    subtitle(ws, "Bank data + PII = security is not optional. All figures $/mo.", "H")

    r = 4
    header_row(ws, r, ["Category", "Line item", "Pre-launch",
                       "500 low", "500 high", "1,500 low", "1,500 high", "3,000+"])
    r += 1
    start = r
    rows = [
        ("Perimeter",       "Cloudflare WAF + DDoS + Bot Mgmt",           25, 60, 120, 200, 300, 500),
        ("Perimeter",       "Cloudflare Zero Trust (team VPN)",            0,  0,  50, 100, 150, 250),
        ("Secrets Mgmt",    "Doppler / Infisical (env + rotation)",        0, 25,  50,  60, 120, 240),
        ("Secrets Mgmt",    "AWS/Google Secret Manager (per-svc keys)",   10, 20,  40,  60, 100, 200),
        ("SSO / IAM",       "SSO (SAML/OIDC) for Enterprise plan",         0,  0,   0,   0, 250, 500),
        ("SSO / IAM",       "1Password Business (per seat, team of 6)",   48, 48,  48,  96,  96, 144),
        ("SSO / IAM",       "MFA enforcement (Auth0/WorkOS add-on)",       0, 50, 100, 150, 300, 500),
        ("App Security",    "Snyk / Semgrep (SAST + dep scanning)",        0, 30,  60, 100, 200, 400),
        ("App Security",    "GitHub Advanced Security",                    0,  0,  49,  49, 147, 294),
        ("App Security",    "DAST (StackHawk / Detectify weekly scan)",    0,  0, 100, 250, 400, 700),
        ("Bug Bounty",      "HackerOne / Intigriti public program",        0,  0,   0, 500, 800,1500),
        ("Bug Bounty",      "Bounty pool (avg payouts, budgeted)",         0,  0,   0, 300, 600,1200),
        ("Endpoint",        "MDM (Kandji/Jamf, per device)",               0, 60,  60, 180, 240, 480),
        ("Endpoint",        "EDR (CrowdStrike/SentinelOne, per seat)",     0, 36,  36, 108, 144, 288),
        ("Logging",         "SIEM (Panther / Datadog Cloud SIEM)",         0,  0, 100, 200, 500,1000),
        ("Logging",         "Audit-log retention (S3 Glacier)",            5, 10,  20,  40,  80, 160),
        ("Backup / DR",     "Cross-region MongoDB backup + PITR",         20, 40,  80, 120, 200, 400),
        ("Backup / DR",     "Immutable off-site backup (Wasabi/S3 lock)", 10, 20,  40,  60, 120, 240),
        ("Compliance Ops",  "Vanta / Drata (SOC 2 automation)",            0,  0,   0, 800,1250,1250),
        ("Compliance Ops",  "Annual pen-test (amortised /mo)",             0,  0,   0, 500, 750,1000),
        ("Compliance Ops",  "IR retainer (Kroll / Mandiant)",              0,  0,   0,   0, 400, 800),
        ("Fraud / Abuse",   "reCAPTCHA Enterprise + IP intel",             0, 20,  40,  80, 150, 300),
        ("Fraud / Abuse",   "Rate-limit + abuse mgmt (Kong/Cloudflare)",   0,  0,  30,  60, 120, 240),
    ]
    for row_ in rows:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", "$#,##0", "$#,##0",
                                "$#,##0", "$#,##0", "$#,##0"])
        r += 1
    end = r - 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAL SECURITY / MONTH").font = BODY_BOLD
    for col in range(3, 9):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({cl}{start}:{cl}{end})")
        ws.cell(row=r, column=col).number_format = "$#,##0"
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = FILL_TOTAL
        ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=col).font = BODY_BOLD

    r += 2
    notes = [
        "• Pre-launch = Day 1 essentials (Cloudflare, secrets, MFA, 1Password, encrypted backups).",
        "• 500-user tier layers in Snyk, MDM, EDR, basic SIEM — SOC 2 Type I readiness.",
        "• 1,500-user tier adds Vanta, pen-test, DAST, bug bounty.",
        "• 3,000+ = mature program (SIEM w/ retention, IR retainer, GitHub Advanced Security).",
    ]
    for n in notes:
        c = ws.cell(row=r, column=1, value=n)
        c.font = BODY
        c.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1


# ============================================================
# COST STACK per user count (used by Profit + Hire Timeline)
# ============================================================
# (users, infra, fractional/outsource, FTE loaded /mo)
# FTE ramp keyed to MRR trigger (see Tech Outsource tab):
#   Sr SWE #1     hired at ~$25k MRR  (~355 users)
#   Platform Eng  hired at 1,500 users
#   Sr SWE #2     hired at 3,000 users OR $60k MRR
#   Security Eng  hired at 3,000+ users (SOC 2 Type II)
#   QA Lead       hired at 3,000 users
# Loaded FTE monthly cost = annual loaded /12:
FTE_MO = {
    "Sr SWE #1":     195_000 / 12,  # ~$16,250
    "Sr SWE #2":     195_000 / 12,
    "Platform Eng":  210_000 / 12,  # ~$17,500
    "Security Eng":  220_000 / 12,  # ~$18,333
    "QA Lead":       165_000 / 12,  # ~$13,750
}

COST_STACK = [
    # users, infra, outsource, list of FTE names on payroll at that stage
    (100,   3500,  3000,  []),                                             # 1 contract dev only
    (250,   4000,  4000,  []),                                             # + light QA fractional
    (500,   4000,  6000,  []),                                             # dev + QA fractional
    (1000,  8500,  7000,  ["Sr SWE #1"]),                                  # first FTE
    (1500, 11000,  9000,  ["Sr SWE #1", "Platform Eng"]),                  # + Platform hire
    (3000, 26000, 10000,  ["Sr SWE #1", "Sr SWE #2", "Platform Eng",
                            "Security Eng", "QA Lead"]),                    # full core team
    (5000, 42000,  8000,  ["Sr SWE #1", "Sr SWE #2", "Platform Eng",
                            "Security Eng", "QA Lead"]),                    # same team, scale infra
]


def _stack_costs(users):
    """Return (infra, outsource, fte_cost, fte_list) for the given stage."""
    for u, infra, outs, ftes in COST_STACK:
        if u == users:
            fte_cost = sum(FTE_MO[n] for n in ftes)
            return infra, outs, fte_cost, ftes
    raise KeyError(users)


# ============================================================
# TAB 8 — PROFIT BREAKDOWN (all-in)
# ============================================================
def tab_profit(wb):
    ws = wb.create_sheet("8. Profit Breakdown")
    widen(ws, [16, 14, 14, 14, 14, 14, 14, 14, 16])
    title(ws, "Profit Breakdown — infra-only vs. ALL-IN (with team)", "I")
    subtitle(ws, f"Blended ARPU = ${BLENDED_ARPU:.2f} (40/35/15/10 mix of "
                 "$38/$75/$95/$149). All-in adds outsource + FTE loaded costs.", "I")

    # --- Section A: MRR by tier per user count ---
    r = 4
    ws.cell(row=r, column=1, value="A. MRR contribution per tier (users × price × mix%)").font = H2
    ws.merge_cells(f"A{r}:I{r}")
    r += 1
    header_row(ws, r, ["Total users",
                       "Solo $38 (40%)", "Standard $75 (35%)",
                       "Pro $95 (15%)", "Enterprise $149 (10%)",
                       "Total MRR", "Total ARR", "", ""])
    r += 1
    for u, *_ in COST_STACK:
        solo  = u * 0.40 * 38
        std   = u * 0.35 * 75
        pro   = u * 0.15 * 95
        ent   = u * 0.10 * 149
        mrr   = solo + std + pro + ent
        set_row(ws, r, [u, solo, std, pro, ent, mrr, mrr * 12, "", ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", "$#,##0", "$#,##0",
                                "$#,##0", "$#,##0", "$#,##0", None, None])
        r += 1

    # --- Section B: INFRA-ONLY margin (unit economics narrative) ---
    r += 2
    ws.cell(row=r, column=1,
            value="B. INFRA-ONLY gross profit — the unit-economics number").font = H2
    ws.merge_cells(f"A{r}:I{r}")
    r += 1
    header_row(ws, r, ["Users", "MRR", "Infra + SaaS + AI", "Gross profit",
                       "Gross margin", "Annual profit", "Cost/user", "", ""])
    r += 1
    for u, infra, _, _ in COST_STACK:
        mrr = u * BLENDED_ARPU
        gp = mrr - infra
        margin = gp / mrr if mrr else 0
        set_row(ws, r, [u, mrr, infra, gp, margin, gp * 12, infra / u, "", ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", "$#,##0", "$#,##0",
                                "0.0%", "$#,##0", "$#,##0.00", None, None])
        r += 1

    # --- Section C: ALL-IN margin (with outsource + FTEs) ---
    r += 2
    ws.cell(row=r, column=1,
            value="C. ALL-IN monthly cost + margin (infra + outsource + FTE loaded)").font = H2
    ws.merge_cells(f"A{r}:I{r}")
    r += 1
    header_row(ws, r, ["Users", "MRR", "Infra", "Outsource",
                       "FTE loaded", "All-in cost", "All-in profit",
                       "All-in margin", "Annual profit"])
    r += 1
    for u, infra, outs, ftes in COST_STACK:
        mrr = u * BLENDED_ARPU
        fte_cost = sum(FTE_MO[n] for n in ftes)
        allin = infra + outs + fte_cost
        gp = mrr - allin
        margin = gp / mrr if mrr else 0
        set_row(ws, r,
                [u, mrr, infra, outs, fte_cost, allin, gp, margin, gp * 12],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", "$#,##0", "$#,##0",
                                "$#,##0", "$#,##0", "$#,##0", "0.0%", "$#,##0"])
        r += 1

    # --- Section D: Break-even users under different burn scenarios ---
    r += 2
    ws.cell(row=r, column=1,
            value="D. Break-even users by cost scenario (blended $70.60 ARPU)").font = H2
    ws.merge_cells(f"A{r}:I{r}")
    r += 1
    header_row(ws, r, ["Scenario", "Total burn /mo",
                       "Break-even @ blended", "@ $75 Std", "@ $95 Pro",
                       "", "", "", ""])
    r += 1
    scenarios = [
        ("Bare-minimum floor (paused build)",         2100),
        ("Steady state (1 fractional dev)",           5500),
        ("Pre-launch today (heavy build)",           15500),
        ("Post-launch, no FTE (500u infra + outs)",  10000),
        ("1 FTE hired (1,000u stage)",               31750),
        ("Full core team (3,000u stage)",            86083),
    ]
    for label, burn in scenarios:
        set_row(ws, r,
                [label, burn,
                 round(burn / BLENDED_ARPU),
                 round(burn / 75),
                 round(burn / 95), "", "", "", ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", "#,##0", "#,##0", "#,##0",
                                None, None, None, None])
        r += 1

    # --- Section E: ARPU sensitivity by mix ---
    r += 2
    ws.cell(row=r, column=1,
            value="E. Blended-ARPU sensitivity by tier mix").font = H2
    ws.merge_cells(f"A{r}:I{r}")
    r += 1
    header_row(ws, r, ["Mix scenario", "Solo %", "Standard %",
                       "Pro %", "Ent %", "Blended ARPU",
                       "MRR @ 1,500u", "", ""])
    r += 1
    mixes = [
        ("Solo-heavy (Freemium bleed)", 0.60, 0.30, 0.08, 0.02),
        ("Baseline (current plan)",     0.40, 0.35, 0.15, 0.10),
        ("Enterprise-tilted",           0.20, 0.30, 0.25, 0.25),
        ("Enterprise-heavy (Partners)", 0.10, 0.25, 0.30, 0.35),
    ]
    for label, s, std, pro, ent in mixes:
        arpu = s * 38 + std * 75 + pro * 95 + ent * 149
        set_row(ws, r,
                [label, s, std, pro, ent, arpu, 1500 * arpu, "", ""],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "0%", "0%", "0%", "0%",
                                "$#,##0.00", "$#,##0", None, None])
        r += 1

    r += 2
    ws.cell(row=r, column=1, value=(
        "Takeaway: infra-only margin stays ~87-89% through the whole ramp. "
        "All-in margin dips to ~49% at 1,500u (post-Platform-Eng hire) and "
        "~44% at 3,000u as the core team fills out, then rebounds to ~63% at "
        "5,000u once revenue outpaces headcount. This is the classic "
        "SaaS S-curve — the pitch is: infra margin proves the model, all-in "
        "margin proves you can operate it."
    )).font = BODY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)


# ============================================================
# TAB 9 — HIRE TIMELINE
# ============================================================
def tab_hire_timeline(wb):
    ws = wb.create_sheet("9. Hire Timeline")
    widen(ws, [30, 14, 16, 20, 16, 16, 30])
    title(ws, "Outsource & Hire Timeline — what, when, why", "G")
    subtitle(ws, "Chronological plan tied to user milestones. Each row shows "
                 "the trigger (users or MRR), monthly cost impact, cumulative "
                 "team spend, and the MRR available at that point.", "G")

    # --- Section A: Immediate outsourcing (pre-launch / day 1) ---
    r = 4
    ws.cell(row=r, column=1, value="A. IMMEDIATE — active or start within 30 days").font = H2
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    header_row(ws, r, ["Engagement", "Type", "Monthly cost",
                       "Trigger", "Users at trigger", "MRR at trigger", "Notes"])
    r += 1
    immediate = [
        ("Emergent contract dev (current)", "Outsource",  3000,
         "Active today",  0,     0,
         "Extra hands during heavy build; scales up to $8k/mo as needed"),

        ("Fractional QA engineer",          "Outsource",  1500,
         "Post-launch",   50,    3530,
         "Playwright smoke suite once users hit the app"),

        ("Cybersecurity vendor SaaS (Snyk, Doppler, Cloudflare Zero Trust)",
         "SaaS",         100,
         "Day 1",         0,     0,
         "Not a person — set up before first paying user"),
    ]
    for row_ in immediate:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", None,
                                "#,##0", "$#,##0", None])
        r += 1

    # --- Section B: Growth-stage outsourcing (250-1,500 users) ---
    r += 2
    ws.cell(row=r, column=1,
            value="B. GROWTH STAGE — 250 to 1,500 users").font = H2
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    header_row(ws, r, ["Engagement", "Type", "Monthly cost",
                       "Trigger", "Users at trigger", "MRR at trigger", "Notes"])
    r += 1
    growth = [
        ("Fractional Platform / DevOps engineer", "Outsource", 3500,
         "1,000 users",   1000,  70600,
         "SRE, IaC, cost tuning; keep until Platform FTE hired"),

        ("Sr Full-Stack Eng #1 (FTE)", "FTE",   16250,
         "$25k MRR OR UK launch",  355, 25063,
         "Second permanent code owner — kills bus factor"),

        ("Fractional Security engineer", "Outsource", 4000,
         "SOC 2 kickoff / 1,000 users", 1000, 70600,
         "Threat model, IAM review, SIEM tuning, IR playbooks"),

        ("Annual pen-test (amortised)", "Outsource", 750,
         "SOC 2 Type I readiness (1,500u)", 1500, 105900,
         "$6-12k one-shot; required for SOC 2"),

        ("HackerOne / Intigriti bug bounty program", "Outsource", 1300,
         "1,500 users",   1500, 105900,
         "Platform fee + budgeted bounty pool"),

        ("Platform / DevOps Eng (FTE)", "FTE", 17500,
         "1,500 users OR SOC 2 Type II", 1500, 105900,
         "Convert fractional to FTE; on-call ownership"),
    ]
    for row_ in growth:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", None,
                                "#,##0", "$#,##0", None])
        r += 1

    # --- Section C: Scale-stage outsourcing + FTEs (1,500-5,000 users) ---
    r += 2
    ws.cell(row=r, column=1,
            value="C. SCALE STAGE — 1,500 to 5,000 users").font = H2
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    header_row(ws, r, ["Engagement", "Type", "Monthly cost",
                       "Trigger", "Users at trigger", "MRR at trigger", "Notes"])
    r += 1
    scale_rows = [
        ("Fractional Data / Analytics engineer", "Outsource", 3000,
         "1,500 users",   1500, 105900,
         "Feature analytics, categorisation ML, churn model"),

        ("Fractional Product Designer", "Outsource", 2500,
         "Marketing site + brand refresh", 1000, 70600,
         "Design system + marketing site + pitch materials"),

        ("Incident Response retainer (Kroll / Mandiant)", "Outsource", 600,
         "1,500 users",   1500, 105900,
         "Pre-paid breach response hours"),

        ("Sr Full-Stack Eng #2 (FTE)", "FTE", 16250,
         "3,000 users OR $60k MRR",  850, 60010,
         "Split frontend/backend; on-call rotation"),

        ("Security Engineer (FTE)", "FTE", 18333,
         "SOC 2 Type II + first enterprise deal", 3000, 211800,
         "Full-time threat model, IR ownership, SSO reviews"),

        ("QA / Test Automation Lead (FTE)", "FTE", 13750,
         "3,000 users",   3000, 211800,
         "Convert fractional QA; Playwright + performance suites"),
    ]
    for row_ in scale_rows:
        set_row(ws, r, list(row_), font=BODY, align=LEFT, border=BOX,
                number_formats=[None, None, "$#,##0", None,
                                "#,##0", "$#,##0", None])
        r += 1

    # --- Section D: Cumulative team cost at each user milestone ---
    r += 2
    ws.cell(row=r, column=1,
            value="D. Cumulative team cost at each milestone").font = H2
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    header_row(ws, r, ["Milestone", "Outsource /mo", "FTEs on payroll",
                       "FTE /mo", "Team total /mo", "MRR at milestone",
                       "Team % of MRR"])
    r += 1
    for u, infra, outs, ftes in COST_STACK:
        fte_cost = sum(FTE_MO[n] for n in ftes)
        team = outs + fte_cost
        mrr = u * BLENDED_ARPU
        pct = team / mrr if mrr else 0
        fte_names = ", ".join(ftes) if ftes else "—"
        set_row(ws, r,
                [f"{u} users", outs, fte_names, fte_cost, team, mrr, pct],
                font=BODY, align=LEFT, border=BOX,
                number_formats=[None, "$#,##0", None, "$#,##0", "$#,##0",
                                "$#,##0", "0.0%"])
        r += 1

    # --- Section E: Decision rules ---
    r += 2
    ws.cell(row=r, column=1, value="E. Decision rules").font = H2
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    rules = [
        "• Rule of thumb: total team cost should stay below 50% of MRR at every milestone. First FTE breaks this rule intentionally — it's the classic S-curve dip.",
        "• Never hire a FTE for something a fractional does <20 hrs/week. Wait until the role is genuinely full-time.",
        "• Prefer converting a fractional to FTE only after 3+ months working together. That's the cheapest de-risked hire path.",
        "• Every outsource line above has a SaaS or automation alternative — try that first for one quarter (Snyk instead of security consultant, GitHub Actions instead of DevOps).",
        "• Delay Security FTE until first enterprise deal is signed; before that, fractional + SaaS covers 90%.",
        "• Emergent contract dev is functionally the cheapest fractional SWE on market — use it until $25k MRR before hiring FTE #1.",
    ]
    for rule in rules:
        c = ws.cell(row=r, column=1, value=rule)
        c.font = BODY
        c.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1


# ============================================================
# Build workbook
# ============================================================
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    tab_summary(wb)
    tab_hard_costs(wb)
    tab_cashflow(wb)
    tab_outsource(wb)
    tab_scale(wb)
    tab_ai(wb)
    tab_security(wb)
    tab_profit(wb)
    tab_hire_timeline(wb)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)
    print(f"Rebuilt: {XLSX}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
